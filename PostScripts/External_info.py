import openpyxl
import re
import socket
import requests
import json
import os
from requests_ntlm import HttpNtlmAuth
import urllib3
import urllib

neb_map = 'neb.map.pre'

url = 'https://itop.severstal.severstalgroup.com/webservices/rest.php?version=1.3'
user = 'noc'
passwd = '1qazBGT%'
ntlm = True

# neb_server = '10.120.63.3'
neb_server = 'localhost'
neb_server_port = '8080'
url_neb_info = "http://" + neb_server + ":" + neb_server_port + "/get?file="+neb_map+"&key=/"
url_neb_list = "http://" + neb_server + ":" + neb_server_port + "/list?file="+neb_map+"&key=/"
# url_neb_info = "http://"+neb_server+":"+neb_server_port+"/get?file=neb.map&key=/area_taparco"
user_neb = "noc"
passwd_neb = "1qaz2wsx"

###########################
###########################
def itop_request(json_data):
    out = {}
    encoded_data = ""
    try:
        encoded_data = json.dumps(json_data)
    except Exception:
        print("Error json to text - " + str(json_data))

    if ntlm:
        response = requests.post(url, data={'json_data': encoded_data}, auth=HttpNtlmAuth(user, passwd), verify=False)
    else:
        response = requests.post(url, data={'auth_user': user, 'auth_pwd': passwd, 'json_data': encoded_data},
                                 verify=False)

    try:
        out = json.loads(response.text)
    except Exception:
        print("Error text to json - " + response.text+"   from json_data - " + str(json_data))
    return out

def get_neb_info():
    result = {}
    if neb_server == 'localhost':
        with open(neb_map, encoding="utf8") as json_file:
            result = json.load(json_file)
    else:
        area_list = requests.get(url_neb_list, headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
        print(area_list.text)
        mas = area_list.text.split("\n")
        for area in mas:
            # if area == "area_perm":
                print("Start - "+area)
                url = url_neb_info+area
                response = requests.get(url, headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
                json_area = json.loads(response.text)
                result[area] = json_area
                print("Stop - " + area)

    return result

def set_extended_info(area, node, info, neb_info):
    url_set = "http://" + neb_server + ":" + neb_server_port + "/set?file=" + neb_map + "&key=/" + area + "/nodes_information/" + node + "/external"
    info['icon'] = 'info.png'
    if neb_info.get(area).get("nodes_information").get(node).get('external'):
        result = requests.post(url_set, urllib.parse.quote_plus(json.dumps({'information': info})),
                               headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
    else:
        result = requests.post(url_set, '{}', headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
        result = requests.post(url_set, urllib.parse.quote_plus(json.dumps({'information': info})),
                               headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
    print(node + ' - ' + json.dumps(info))
###########################
###########################

ip_info = {}
mac_info = {}
sysname_info = {}

print("Starting get Neb node info ...")
neb_info = get_neb_info()
print("Stop get Neb node info.")

##################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Костомукша\\KARO Active.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Костомукша\\KARO Active.xlsx')
sheet = wb['devices']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['D'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['C' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            description = sheet['E' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            sw = sheet['A' + str(i)].value
            if sw:
                sw = str(sw).strip().replace("\n", "")
                info['sw'] = sw
            else:
                sw = ''
            # date = sheet['P' + str(i)].value
            # if date:
            #     date = str(date)
            #     info['date'] = date
            # else:
            #     date = ''
            model = sheet['F' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['H' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['G' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)

##################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Колпино\\Kolpino.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Колпино\\Kolpino.xlsx')
sheet = wb['Devices']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['D'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['A' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            description = sheet['G' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            # date = sheet['P' + str(i)].value
            # if date:
            #     date = str(date)
            #     info['date'] = date
            # else:
            #     date = ''
            model = sheet['F' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['H' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['I' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)

##################################################################################
print('Get info from \\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\СС-Инвест\\Latvia_and_Poland\\Latvia\\Latvia Switches.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\СС-Инвест\\Latvia_and_Poland\\Latvia\\Latvia Switches.xlsx')
sheet = wb['Devices']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['B'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['C' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            description = sheet['D' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            # date = sheet['P' + str(i)].value
            # if date:
            #     date = str(date)
            #     info['date'] = date
            # else:
            #     date = ''
            model = sheet['F' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['G' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['H' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)

##################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Гипрошахт\\Гипрошахт Active.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Гипрошахт\\Гипрошахт Active.xlsx')
sheet = wb['Active']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['A'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['D' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            description = sheet['F' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            date = sheet['P' + str(i)].value
            if date:
                date = str(date)
                info['date'] = date
            else:
                date = ''
            model = sheet['H' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['I' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['J' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)
##################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Олкон\\Olkon Active.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Олкон\\Olcon Active.xlsx')
sheet = wb['Active']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['A'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['D' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            description = sheet['E' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            date = sheet['O' + str(i)].value
            if date:
                date = str(date)
                info['date'] = date
            else:
                date = ''
            model = sheet['G' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['H' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['I' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)
##################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Воркута\\АСО_Воркута.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Предприятия\\Воркута\\АСО_Воркута.xlsx')
sheet = wb['ACO']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['A'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sysname = sheet['C' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            date = sheet['I' + str(i)].value
            if date:
                date = str(date)
                info['date'] = date
            else:
                date = ''
            model = sheet['E' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['F' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''
            mac = sheet['G' + str(i)].value
            if mac:
                mac = str(mac).strip().replace("\n", "")
                info['mac'] = mac
            else: mac = ''
            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)
#############################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Учет ресурсов\\ЧерМК Active.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\CorporateInf\\IT\\NOC\\Cher\\NOC CHER\\Учет ресурсов\\ЧерМК Active.xlsx')
sheet = wb['Active']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['C'+str(i)].value
    if val:
        val = val.strip()
        if re.match(r'^\d+\.\d+\.\d+\.\d+$', val):
            ip = val
            info['ip'] = ip
            sw = sheet['D' + str(i)].value
            if sw:
                sw = str(sw).strip().replace("\n", "")
                info['sw'] = sw
            else:
                sw = ''
            sysname = sheet['E' + str(i)].value
            if sysname:
                sysname = str(sysname).strip().replace("\n", "")
                info['sysname'] = sysname
            else:
                sysname = ''
            ceh = sheet['F' + str(i)].value
            if ceh:
                ceh = str(ceh).strip().replace("\n", "")
                info['ceh'] = ceh
            else:
                ceh = ''
            description = sheet['G' + str(i)].value
            if description:
                description = str(description).strip().replace("\n", "")
                info['description'] = description
            else:
                description = ''
            date = sheet['H' + str(i)].value
            if date:
                date = str(date)
                info['date'] = date
            else:
                date = ''
            position = sheet['I' + str(i)].value
            if position:
                position = str(position).strip().replace("\n", "")
                info['position'] = position
            else:
                position = ''
            tsh = sheet['K' + str(i)].value
            if tsh:
                tsh = str(tsh).strip().replace("\n", "")
                info['tsh'] = tsh
            else:
                tsh = ''
            model = sheet['M' + str(i)].value
            if model:
                model = str(model).strip().replace("\n", "")
                info['model'] = model
            else:
                model = ''
            serial = sheet['N' + str(i)].value
            if serial:
                serial = str(serial).strip().replace("\n", "")
                info['serial'] = serial
            else: serial = ''

            if sysname:
                sysname_info[sysname.lower()] = info
            if ip:
                ip_info[ip] = info
                # print(ip+"- \t"+sw+"\t"+sysname+"\t"+ceh+"\t"+description+"\t"+date+"\t"+position+"\t"+tsh+"\t"+model+"\t"+serial)

###############################################################################################################
print('Get info from \\\\workspace.severstal.com@ssl\\project_office\\Infocom\\ArchPr\\pr2014\\Wi-Fi\\WiFi2\\Список точек доступа.xlsx')
wb = openpyxl.load_workbook('\\\\workspace.severstal.com@ssl\\project_office\\Infocom\\ArchPr\\pr2014\\Wi-Fi\\WiFi2\\Список точек доступа.xlsx')
sheet = wb['Список точек доступа']
rows = sheet.max_row
# cols = sheet.max_column

for i in range(1, rows):
    info = {}
    val = sheet['D'+str(i)].value
    if val:
        val = val.strip()
        sysname = val
        info['sysname'] = sysname

        ip = ''
        try:
            ip = socket.gethostbyname(sysname)
        except:
            pass

        if ip and re.match(r'^\d+\.\d+\.\d+\.\d+$', ip):
            info['ip'] = ip
        model = sheet['E' + str(i)].value
        if model:
            model = str(model).strip().replace("\n", "")
            info['sw'] = model
        else:
            model = ''
        antena = sheet['F' + str(i)].value
        if antena:
            antena = str(antena).strip().replace("\n", "")
            info['antena'] = antena
        else:
            antena = ''
        serial = sheet['G' + str(i)].value
        if serial:
            serial = str(serial).strip().replace("\n", "")
            info['serial'] = serial
        else:
            serial = ''
        mac = sheet['H' + str(i)].value
        if mac:
            mac = str(mac).strip().replace("\n", "")
            info['mac'] = mac
        else:
            mac = ''
        description = sheet['J' + str(i)].value
        if description:
            description = str(description).strip().replace("\n", "")
            info['description'] = description
        else:
            description = ''
        date = sheet['P' + str(i)].value
        if date:
            date = str(date)
            info['date'] = date
        else:
            date = ''
        position = sheet['Q' + str(i)].value
        if position:
            position = str(position).strip().replace("\n", "")
            info['position'] = position
        else:
            position = ''

        if sysname:
            if ip:
                ip_info[ip] = info
            if mac:
                mac_info[mac.replace(':', '').replace('.', '').lower()] = info
            sysname_info[sysname.lower()] = info
            # print(ip+"- \t"+sysname+"\t"+description+"\t"+date+"\t"+position+"\t"+"\t"+model+"\t"+serial)

####################################################################################
urllib3.disable_warnings()
json_data = {
    "operation": "core/get",
    "class": "IPCamera",
    "key": "SELECT IPCamera",
    # "key": "SELECT IPCamera WHERE name = 'Agp-agc2-K-1-3-0-10'",
    "output_fields": "name, move2production, project, serialnumber, brand_name, model_name, macaddress, cameravsmpart, usercomment, managementip_id_friendlyname"
    # "output_fields": "*"
}

result = itop_request(json_data)

if result.get('objects'):
    for item in result.get('objects'):
        info = {}
        ip = ''
        if result.get('objects').get(item).get('fields').get('managementip_id_friendlyname'):
            ip = result.get('objects').get(item).get('fields').get('managementip_id_friendlyname')
            info['ip'] = ip.strip().replace("\n", " ")
        sysname = ''
        if result.get('objects').get(item).get('fields').get('name'):
            sysname = result.get('objects').get(item).get('fields').get('name')
            info['sysname'] = sysname.strip().replace("\n", " ")
        date = ''
        if result.get('objects').get(item).get('fields').get('move2production'):
            date = result.get('objects').get(item).get('fields').get('move2production')
            info['date'] = date.strip().replace("\n", " ")
        project = ''
        if result.get('objects').get(item).get('fields').get('project'):
            project = result.get('objects').get(item).get('fields').get('project')
            info['project'] = project.strip().replace("\n", " ")
        serial = ''
        if result.get('objects').get(item).get('fields').get('serialnumber'):
            serial = result.get('objects').get(item).get('fields').get('serialnumber')
            info['serial'] = serial.strip().replace("\n", " ")
        brand = ''
        if result.get('objects').get(item).get('fields').get('brand_name'):
            brand = result.get('objects').get(item).get('fields').get('brand_name')
            info['brand'] = brand.strip().replace("\n", " ")
        model = ''
        if result.get('objects').get(item).get('fields').get('model_name'):
            model = result.get('objects').get(item).get('fields').get('model_name')
            info['model'] = model.strip().replace("\n", " ")
        mac = ''
        if result.get('objects').get(item).get('fields').get('macaddress'):
            mac = result.get('objects').get(item).get('fields').get('macaddress')
            info['mac'] = mac.strip().replace("\n", " ")
        camera_vsm_part = ''
        if result.get('objects').get(item).get('fields').get('cameravsmpart'):
            camera_vsm_part = result.get('objects').get(item).get('fields').get('cameravsmpart')
            info['camera_vsm_part'] = camera_vsm_part.strip().replace("\n", " ")
        description = ''
        if result.get('objects').get(item).get('fields').get('usercomment'):
            description = result.get('objects').get(item).get('fields').get('usercomment')
            info['description'] = description.strip().replace("\n", " ")

        if ip:
            ip_info[ip] = info
        if mac:
            mac_info[mac.replace(':', '').replace('.', '').lower()] = info
        if sysname:
            sysname_info[sysname.lower()] = info
        # print(ip + "- \t" + sysname + "\t" + description + "\t" + date + "\t" + "\t" + model + "\t" + serial)
####################################################################################

urllib3.disable_warnings()
json_data = {
    "operation": "core/get",
    "class": "NetworkDevice",
    "key": "SELECT NetworkDevice",
    # "key": "SELECT NetworkDevice WHERE name = 'cs2950-24_volg_abk_kc-2_l1_sortirovka_n1'",
    "output_fields": "name, macaddress, managementip_name, description, mod_date, asset_number"
    # "output_fields": "*"
}

result = itop_request(json_data)

if result.get('objects'):
    for item in result.get('objects'):
        info = {}

        sysname = ''
        if result.get('objects').get(item).get('fields').get('name'):
            sysname = result.get('objects').get(item).get('fields').get('name')
            info['sysname'] = sysname.strip().replace("\n", " ")
        mac = ''
        if result.get('objects').get(item).get('fields').get('macaddress'):
            mac = result.get('objects').get(item).get('fields').get('macaddress')
            info['mac'] = mac.strip().replace("\n", " ")
        ip = ''
        if result.get('objects').get(item).get('fields').get('managementip_name'):
            ip = result.get('objects').get(item).get('fields').get('managementip_name')
            info['ip'] = ip.strip().replace("\n", " ")
        date = ''
        if result.get('objects').get(item).get('fields').get('mod_date'):
            date = result.get('objects').get(item).get('fields').get('mod_date')
            info['date'] = date.strip().replace("\n", " ")
        description = ''
        if result.get('objects').get(item).get('fields').get('description'):
            description = result.get('objects').get(item).get('fields').get('description')
            info['description'] = description.strip().replace("\n", " ")
        sw = ''
        if result.get('objects').get(item).get('fields').get('asset_number'):
            sw = result.get('objects').get(item).get('fields').get('asset_number')
            info['sw'] = sw.strip().replace("\n", " ")

        if ip:
            ip_info[ip] = info
        if mac:
            mac_info[mac.replace(':', '').replace('.', '').lower()] = info
        if sysname:
            sysname_info[sysname.lower()] = info
        # if description:
            # print(ip + "- \t" + sysname + "\t" + description + "\t" + date + "\t" + "\t" + mac + "\t" + ip)

####################################################################################
for area in neb_info:
    if neb_info.get(area) and neb_info.get(area).get("nodes_information"):
        for node in neb_info.get(area).get("nodes_information"):
            if neb_info.get(area).get("nodes_information").get(node).get("general"):
                mac = neb_info.get(area).get("nodes_information").get(node).get("general").get("base_address")
                if mac:
                    mac = mac.replace(':', '').replace('.', '').lower()
            if neb_info.get(area).get("nodes_information").get(node).get("general"):
                sysname = neb_info.get(area).get("nodes_information").get(node).get("general").get("sysname")
                if sysname:
                    sysname = sysname.lower()
            if ip_info.get(node):
                set_extended_info(area, node, ip_info[node], neb_info)
            elif mac and mac_info.get(mac):
                set_extended_info(area, node, mac_info[mac], neb_info)
            elif sysname and sysname_info.get(sysname):
                ip = sysname_info[sysname].get('ip')
                if ip:
                    if ip == node:
                        set_extended_info(area, node, sysname_info[sysname], neb_info)
                else:
                    set_extended_info(area, node, sysname_info[sysname], neb_info)
            # else:
            #     print(node+" is not information!")

commit_url = "http://" + neb_server + ":" + neb_server_port + "/commit"
res = requests.get(commit_url, headers={"user": user_neb, "passwd": passwd_neb}, verify=False)
# print("1111111")
